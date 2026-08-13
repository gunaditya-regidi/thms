import os
import csv
import io
import uuid
import zipfile
from datetime import datetime
from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file, current_app, jsonify
from flask_login import login_required, current_user
from models import db, User, Team, Member, House, Task, TaskCompletion, Round1Approval, QRCode, Round2Progress, QRScanLog, SystemLog, Manager
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash

# ReportLab imports for PDF generation
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

# OpenPyXL imports for Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

admin_bp = Blueprint('admin', __name__)

def check_admin_role(func):
    """Decorator to ensure user is logged in as Admin."""
    def wrapper(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash("Unauthorized. Please log in as an Administrator.", "danger")
            return redirect(url_for('auth.login'))
        return func(*args, **kwargs)
    wrapper.__name__ = func.__name__
    return wrapper

def generate_qr_image(uuid_str):
    """Generates a QR code image pointing to the direct scan URL."""
    import qrcode
    qr_dir = current_app.config['QR_FOLDER']
    os.makedirs(qr_dir, exist_ok=True)
    
    # We construct the URL dynamically
    scan_url = url_for('team.direct_scan', uuid=uuid_str, _external=True)
    
    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(scan_url)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    filename = f"qr_{uuid_str}.png"
    filepath = os.path.join(qr_dir, filename)
    img.save(filepath)
    return filename

@admin_bp.route('/admin')
@check_admin_role
def dashboard():
    # Gather general statistics
    total_teams = Team.query.count()
    total_managers = Manager.query.count()
    total_tasks = Task.query.count()
    total_qrs = QRCode.query.count()
    
    # Houses summary
    houses = House.query.all()
    
    # Round progress counts
    r1_pending_start = Team.query.filter_by(round1_status='pending_start').count()
    r1_active = Team.query.filter_by(current_round=1, round1_status='active').count()
    r1_pending_approval = Team.query.filter_by(round1_status='requested').count()
    r1_verified = Team.query.filter_by(round1_status='verified').count()
    r2_active = Team.query.filter_by(current_round=2, round2_completed=False).count()
    r2_completed = Team.query.filter_by(round2_completed=True).count()
    
    # Fetch recent registrations and scan activity
    recent_teams = Team.query.order_by(Team.created_at.desc()).limit(5).all()
    recent_scans = QRScanLog.query.order_by(QRScanLog.timestamp.desc()).limit(5).all()

    users = User.query.all()
    return render_template('admin/dashboard.html',
                           total_teams=total_teams,
                           total_managers=total_managers,
                           total_tasks=total_tasks,
                           total_qrs=total_qrs,
                           houses=houses,
                           r1_pending_start=r1_pending_start,
                           r1_active=r1_active,
                           r1_pending_approval=r1_pending_approval,
                           r1_verified=r1_verified,
                           r2_active=r2_active,
                           r2_completed=r2_completed,
                           recent_teams=recent_teams,
                           recent_scans=recent_scans,
                           users=users)

@admin_bp.route('/admin/edit-manager/<int:user_id>', methods=['POST'])
@check_admin_role
def edit_manager(user_id):
    from models import User
    user = User.query.get_or_404(user_id)
    if user.role != 'manager':
        flash("User is not a manager.", "danger")
        return redirect(url_for('admin.dashboard'))
        
    new_username = request.form.get('username', '').strip()
    new_password = request.form.get('password', '').strip()
    
    if not new_username:
        flash("Manager name cannot be empty.", "danger")
        return redirect(url_for('admin.dashboard'))
        
    # Check if username is already taken by another user
    existing = User.query.filter(User.username == new_username, User.id != user_id).first()
    if existing:
        flash(f"The name '{new_username}' is already taken.", "danger")
        return redirect(url_for('admin.dashboard'))
        
    old_username = user.username
    user.username = new_username
    if new_password:
        user.set_password(new_password)
        details = f"Admin updated manager '{old_username}' to name '{new_username}' and updated password."
    else:
        details = f"Admin updated manager '{old_username}' name to '{new_username}'."
        
    # Log the action
    log = SystemLog(
        action='edit_manager',
        details=details,
        user_id=current_user.id,
        ip_address=request.remote_addr,
        browser=request.user_agent.string
    )
    db.session.add(log)
    db.session.commit()
    
    # Broadcast to all clients to update dashboards
    socketio = current_app.extensions.get('socketio')
    if socketio:
        socketio.emit('manager_updated', {
            'manager_id': user.id,
            'new_username': new_username
        })
        
    flash(f"Manager '{new_username}' updated successfully!", "success")
    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/admin/bulk-teams-action', methods=['POST'])
@check_admin_role
def bulk_teams_action():
    data = request.get_json() or {}
    action = data.get('action')
    team_ids = data.get('team_ids', [])
    
    if not action or not team_ids:
        return jsonify({'status': 'error', 'message': 'Missing action or team IDs.'}), 400
        
    from models import Team, User, db, SystemLog
    
    try:
        teams = Team.query.filter(Team.id.in_(team_ids)).all()
        if not teams:
            return jsonify({'status': 'error', 'message': 'No teams found for the provided IDs.'}), 404
            
        user_ids = [t.user_id for t in teams]
        users = User.query.filter(User.id.in_(user_ids)).all()
        
        count = 0
        if action == 'suspend':
            for u in users:
                u.is_active = False
            details = f"Admin suspended teams: {', '.join([t.team_name for t in teams])}"
            count = len(users)
        elif action == 'unsuspend':
            for u in users:
                u.is_active = True
            details = f"Admin unsuspended teams: {', '.join([t.team_name for t in teams])}"
            count = len(users)
        elif action == 'delete':
            for u in users:
                db.session.delete(u)
            details = f"Admin deleted teams: {', '.join([t.team_name for t in teams])}"
            count = len(users)
        else:
            return jsonify({'status': 'error', 'message': 'Invalid action.'}), 400
            
        log = SystemLog(
            action=f'bulk_{action}',
            details=details,
            user_id=current_user.id,
            ip_address=request.remote_addr,
            browser=request.user_agent.string
        )
        db.session.add(log)
        db.session.commit()
        
        socketio = current_app.extensions.get('socketio')
        if socketio:
            socketio.emit('team_reset', {})
            
        return jsonify({'status': 'success', 'message': f'Successfully performed {action} on {count} teams.'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@admin_bp.route('/admin/teams')
@check_admin_role
def teams_list():
    import datetime
    teams = Team.query.order_by(Team.created_at.desc()).all()
    server_now_iso = datetime.datetime.utcnow().isoformat() + "Z"
    return render_template('admin/teams.html', teams=teams, server_now_iso=server_now_iso)

@admin_bp.route('/admin/approve-r1/<int:team_id>', methods=['POST'])
@check_admin_role
def approve_r1(team_id):
    team = Team.query.get_or_404(team_id)
    if team.round1_status not in ['requested', 'verified']:
        flash("Round 1 completion has not been requested or is already approved.", "warning")
        return redirect(url_for('admin.teams_list'))

    try:
        team.round1_status = 'approved'
        team.current_round = 2 # Automatically unlocks Round 2!
        team.round2_current_clue = 2
        
        # Load or create approval record
        approval = Round1Approval.query.filter_by(team_id=team_id).first()
        if not approval:
            approval = Round1Approval(team_id=team_id)
            db.session.add(approval)
            
        approval.approved_at = datetime.utcnow()
        approval.approved_by_admin_id = current_user.id

        # Audit log
        log = SystemLog(
            action='r1_approved',
            details=f"Admin '{current_user.username}' approved Round 1 for Team '{team.team_name}'. Official completion timestamp saved.",
            user_id=current_user.id,
            team_id=team_id,
            ip_address=request.remote_addr,
            browser=request.user_agent.string
        )
        db.session.add(log)
        db.session.commit()

        # Emit SocketIO event
        socketio = current_app.extensions.get('socketio')
        if socketio:
            socketio.emit('r1_approved', {
                'team_id': team_id,
                'team_name': team.team_name,
                'house': team.house.name,
                'timestamp': approval.approved_at.strftime('%Y-%m-%d %H:%M:%S')
            })

        flash(f"Round 1 approved and Round 2 unlocked for Team {team.team_name}.", "success")
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error in approving R1 completion: {str(e)}")
        flash("An error occurred. Please try again.", "danger")

    return redirect(url_for('admin.teams_list'))

@admin_bp.route('/admin/approve-start/<int:team_id>', methods=['POST'])
@check_admin_role
def approve_start(team_id):
    team = Team.query.get_or_404(team_id)
    if team.round1_status != 'pending_start':
        flash("Team is already started or past starting phase.", "warning")
        return redirect(url_for('admin.teams_list'))

    try:
        from datetime import datetime
        now = datetime.utcnow()
        team.round1_status = 'active'
        team.created_at = now
        
        log = SystemLog(
            action='r1_started',
            details=f"Admin '{current_user.username}' approved Team '{team.team_name}' to start Round 1.",
            user_id=current_user.id,
            team_id=team_id,
            ip_address=request.remote_addr,
            browser=request.user_agent.string
        )
        db.session.add(log)
        db.session.commit()

        socketio = current_app.extensions.get('socketio')
        if socketio:
            socketio.emit('team_started', {
                'team_id': team_id,
                'team_name': team.team_name,
                'house': team.house.name,
                'timestamp': now.strftime('%Y-%m-%d %H:%M:%S')
            })

        flash(f"Team {team.team_name} approved to start Round 1.", "success")
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error in starting team: {str(e)}")
        flash("An error occurred. Please try again.", "danger")

    return redirect(url_for('admin.teams_list'))

@admin_bp.route('/admin/approve-start-all', methods=['POST'])
@check_admin_role
def approve_start_all():
    pending_teams = Team.query.filter_by(round1_status='pending_start').all()
    if not pending_teams:
        flash("No teams are currently pending approval to start.", "warning")
        return redirect(url_for('admin.dashboard'))

    try:
        from datetime import datetime
        now = datetime.utcnow()
        socketio = current_app.extensions.get('socketio')
        
        count = 0
        for team in pending_teams:
            team.round1_status = 'active'
            team.created_at = now
            
            log = SystemLog(
                action='r1_started',
                details=f"Admin '{current_user.username}' approved Team '{team.team_name}' to start Round 1 via universal start.",
                user_id=current_user.id,
                team_id=team.id,
                ip_address=request.remote_addr,
                browser=request.user_agent.string
            )
            db.session.add(log)
            count += 1
            
            if socketio:
                socketio.emit('team_started', {
                    'team_id': team.id,
                    'team_name': team.team_name,
                    'house': team.house.name,
                    'timestamp': now.strftime('%Y-%m-%d %H:%M:%S')
                })

        db.session.commit()
        flash(f"Successfully approved and started {count} teams simultaneously!", "success")
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error in starting all teams: {str(e)}")
        flash("An error occurred while starting all teams.", "danger")

    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/admin/declare-winner/<int:team_id>', methods=['POST'])
@check_admin_role
def declare_winner(team_id):
    team = Team.query.get_or_404(team_id)
    if not team.round2_completed:
        flash("This team has not completed Round 2/Clue 7 yet.", "warning")
        return redirect(url_for('admin.teams_list'))
        
    rank_val = request.form.get('winner_rank', '').strip()
    if rank_val == '':
        team.winner_rank = None
        db.session.commit()
        flash(f"Removed winner rank declaration for {team.team_name}.", "info")
    else:
        try:
            rank = int(rank_val)
            if rank not in [1, 2, 3]:
                flash("Invalid winner rank specified.", "danger")
                return redirect(url_for('admin.teams_list'))
                
            # Clear other team that has this rank assigned
            existing = Team.query.filter_by(winner_rank=rank).first()
            if existing and existing.id != team.id:
                existing.winner_rank = None
                
            team.winner_rank = rank
            db.session.add(team)
            db.session.commit()
            
            # Emit Socket.IO event to update stats dynamically
            socketio = current_app.extensions.get('socketio')
            if socketio:
                socketio.emit('winners_updated', {})
                
            place_names = {1: "🥇 1st Place", 2: "🥈 2nd Place", 3: "🥉 3rd Place"}
            flash(f"Successfully declared {team.team_name} as {place_names[rank]} Winner!", "success")
        except ValueError:
            flash("Invalid rank format.", "danger")
            
    return redirect(url_for('admin.teams_list'))

@admin_bp.route('/admin/promote-r2/<int:team_id>', methods=['POST'])
@check_admin_role
def promote_r2(team_id):
    team = Team.query.get_or_404(team_id)
    try:
        # Complete all 10 tasks if they aren't complete
        tasks = Task.query.all()
        now = datetime.utcnow()
        for task in tasks:
            existing = TaskCompletion.query.filter_by(team_id=team.id, task_id=task.id).first()
            if not existing:
                house_manager = Manager.query.filter_by(house_id=team.house_id).first()
                manager_id = house_manager.id if house_manager else 1
                
                tc = TaskCompletion(
                    team_id=team.id,
                    task_id=task.id,
                    completed_at=now,
                    verified_by_manager_id=manager_id
                )
                db.session.add(tc)

        # Set status approved and round to 2
        team.round1_status = 'approved'
        team.current_round = 2
        team.round2_current_clue = 2
        
        # Build approval record
        approval = Round1Approval.query.filter_by(team_id=team.id).first()
        if not approval:
            approval = Round1Approval(team_id=team.id)
            db.session.add(approval)
        
        if not approval.requested_at:
            approval.requested_at = now
        if not approval.verified_at:
            approval.verified_at = now
        approval.approved_at = now
        approval.approved_by_admin_id = current_user.id
        
        # Audit log
        log = SystemLog(
            action='team_promoted_r2',
            details=f"Admin '{current_user.username}' directly PROMOTED Team '{team.team_name}' to Round 2.",
            user_id=current_user.id,
            team_id=team.id,
            ip_address=request.remote_addr,
            browser=request.user_agent.string
        )
        db.session.add(log)
        db.session.commit()

        # Emit SocketIO event
        socketio = current_app.extensions.get('socketio')
        if socketio:
            socketio.emit('r1_approved', {
                'team_id': team.id,
                'team_name': team.team_name,
                'house': team.house.name,
                'timestamp': now.strftime('%Y-%m-%d %H:%M:%S')
            })

        flash(f"Team {team.team_name} successfully promoted directly to Round 2.", "success")
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error promoting team: {str(e)}")
        flash("An error occurred during promotion.", "danger")

    return redirect(url_for('admin.teams_list'))


@admin_bp.route('/admin/reset/<int:team_id>', methods=['POST'])
@check_admin_role
def reset_team(team_id):
    team = Team.query.get_or_404(team_id)
    try:
        # Clear completions, approvals, and round 2 progress
        TaskCompletion.query.filter_by(team_id=team_id).delete()
        Round1Approval.query.filter_by(team_id=team_id).delete()
        Round2Progress.query.filter_by(team_id=team_id).delete()
        
        # Reset team attributes
        team.current_round = 1
        team.round1_status = 'pending_start'
        team.round2_current_clue = 2
        team.round2_completed = False
        team.round2_completion_time = None
        
        # Audit log
        log = SystemLog(
            action='team_reset',
            details=f"Admin '{current_user.username}' reset all progress (R1 & R2) for Team '{team.team_name}'.",
            user_id=current_user.id,
            team_id=team_id,
            ip_address=request.remote_addr,
            browser=request.user_agent.string
        )
        db.session.add(log)
        db.session.commit()

        # Emit SocketIO event
        socketio = current_app.extensions.get('socketio')
        if socketio:
            socketio.emit('team_reset', {
                'team_id': team_id,
                'team_name': team.team_name
            })

        flash(f"Team {team.team_name} progress has been completely reset.", "success")
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error resetting team: {str(e)}")
        flash("An error occurred during reset.", "danger")

    return redirect(url_for('admin.teams_list'))

@admin_bp.route('/admin/reset-qr/<int:team_id>', methods=['POST'])
@check_admin_role
def reset_qr_progress(team_id):
    team = Team.query.get_or_404(team_id)
    try:
        # Clear round 2 progress only
        Round2Progress.query.filter_by(team_id=team_id).delete()
        
        team.round2_current_clue = 2
        team.round2_completed = False
        team.round2_completion_time = None
        
        # Audit log
        log = SystemLog(
            action='team_qr_reset',
            details=f"Admin '{current_user.username}' reset Round 2 QR clue progress for Team '{team.team_name}'.",
            user_id=current_user.id,
            team_id=team_id,
            ip_address=request.remote_addr,
            browser=request.user_agent.string
        )
        db.session.add(log)
        db.session.commit()

        # Emit SocketIO
        socketio = current_app.extensions.get('socketio')
        if socketio:
            socketio.emit('team_qr_reset', {
                'team_id': team_id,
                'team_name': team.team_name
            })

        flash(f"Round 2 QR progress has been reset for Team {team.team_name}.", "success")
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error resetting QR progress: {str(e)}")
        flash("An error occurred.", "danger")

    return redirect(url_for('admin.teams_list'))

@admin_bp.route('/admin/reset-all', methods=['POST'])
@check_admin_role
def reset_all_teams():
    try:
        # Delete task completions, approvals, progresses, and scan logs
        TaskCompletion.query.delete()
        Round1Approval.query.delete()
        Round2Progress.query.delete()
        QRScanLog.query.delete()
        
        # Reset all team status values
        teams = Team.query.all()
        for team in teams:
            team.current_round = 1
            team.round1_status = 'pending_start'
            team.round2_current_clue = 2
            team.round2_completed = False
            team.round2_completion_time = None
            
        # Log system event
        sys_log = SystemLog(
            action='system_reset_all',
            details=f"Admin '{current_user.username}' reset ALL teams and clue progress to zero.",
            user_id=current_user.id,
            ip_address=request.remote_addr,
            browser=request.user_agent.string
        )
        db.session.add(sys_log)
        db.session.commit()
        
        # Emit SocketIO global reset signal so all dashboards reload/update
        socketio = current_app.extensions.get('socketio')
        if socketio:
            socketio.emit('global_reset', {})
            
        flash("All teams, tasks, and clue progress have been completely reset to zero.", "success")
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error executing global reset: {str(e)}")
        flash("An error occurred during global reset.", "danger")
        
    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/admin/qrs', methods=['GET', 'POST'])
@check_admin_role
def manage_qrs():
    if request.method == 'POST':
        is_dummy = request.form.get('is_dummy') == 'true'
        clue_number = int(request.form.get('clue_number', 0)) if not is_dummy else 99
        password = request.form.get('password', '').strip() if not is_dummy else 'dummy-decoy'
        hint = 'Decoy Clue'
        
        if not is_dummy and not password:
            flash("Password is required for standard clues.", "danger")
            return redirect(url_for('admin.manage_qrs'))

        # Image is required for dummy decoy clues
        image_file = request.files.get('image')
        if is_dummy and (not image_file or image_file.filename == ''):
            flash("Image upload is required for decoy dummy clues.", "danger")
            return redirect(url_for('admin.manage_qrs'))

        try:
            # Handle image upload
            image_path = None
            image_base64 = None
            if 'image' in request.files:
                file = request.files['image']
                if file and file.filename != '':
                    filename = secure_filename(f"clue_{clue_number}_{uuid.uuid4().hex[:6]}_{file.filename}")
                    upload_dir = current_app.config['UPLOAD_FOLDER']
                    os.makedirs(upload_dir, exist_ok=True)
                    file_save_path = os.path.join(upload_dir, filename)
                    
                    import base64
                    file_data = file.read()
                    file.seek(0)
                    image_base64 = base64.b64encode(file_data).decode('utf-8')
                    
                    file.save(file_save_path)
                    image_path = f"/static/uploads/{filename}"

            # Process allowed houses list
            allowed_houses_list = request.form.getlist('allowed_houses[]')
            allowed_houses = ",".join(allowed_houses_list) if allowed_houses_list else None

            # Create code record
            code_uuid = str(uuid.uuid4())
            new_qr = QRCode(
                uuid=code_uuid,
                clue_number=clue_number if not is_dummy else 99,
                password=password,
                hint=hint,
                image_path=image_path,
                image_base64=image_base64,
                is_dummy=is_dummy,
                allowed_houses=allowed_houses
            )
            db.session.add(new_qr)
            db.session.flush() # gets ID for filename creation
            
            # Generate QR code PNG (except for Clue 1)
            if clue_number != 1 or is_dummy:
                generate_qr_image(code_uuid)

            # System log
            log = SystemLog(
                action='qr_generated',
                details=f"Admin '{current_user.username}' generated {'Dummy ' if is_dummy else ''}QR Clue {clue_number} (UUID: {code_uuid}).",
                user_id=current_user.id
            )
            db.session.add(log)
            db.session.commit()

            flash("QR Code generated successfully!", "success")
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error generating QR code: {str(e)}")
            flash("An error occurred while generating QR Code.", "danger")
            
        return redirect(url_for('admin.manage_qrs'))

    qrs = QRCode.query.filter(QRCode.clue_number <= 7).order_by(QRCode.is_dummy, QRCode.clue_number).all()
    return render_template('admin/qrs.html', qrs=qrs)

@admin_bp.route('/admin/qrs/delete/<int:qr_id>', methods=['POST'])
@check_admin_role
def delete_qr(qr_id):
    qr = QRCode.query.get_or_404(qr_id)
    try:
        # Delete file from directory if exists
        filename = f"qr_{qr.uuid}.png"
        filepath = os.path.join(current_app.config['QR_FOLDER'], filename)
        if os.path.exists(filepath):
            os.remove(filepath)
            
        # Delete clue image if uploaded
        if qr.image_path:
            img_filename = qr.image_path.split('/')[-1]
            img_filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], img_filename)
            if os.path.exists(img_filepath):
                os.remove(img_filepath)

        db.session.delete(qr)
        
        # Log action
        log = SystemLog(
            action='qr_deleted',
            details=f"Admin '{current_user.username}' deleted QR Clue {qr.clue_number} (UUID: {qr.uuid}).",
            user_id=current_user.id
        )
        db.session.add(log)
        db.session.commit()
        
        flash("QR Code deleted successfully.", "success")
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting QR: {str(e)}")
        flash("An error occurred during deletion.", "danger")

    return redirect(url_for('admin.manage_qrs'))

@admin_bp.route('/admin/qrs/delete-all', methods=['POST'])
@check_admin_role
def delete_all_qrs():
    qrs = QRCode.query.all()
    deleted_count = 0
    try:
        for qr in qrs:
            filename = f"qr_{qr.uuid}.png"
            filepath = os.path.join(current_app.config['QR_FOLDER'], filename)
            if os.path.exists(filepath):
                try:
                    os.remove(filepath)
                except Exception:
                    pass
                
            if qr.image_path:
                img_filename = qr.image_path.split('/')[-1]
                img_filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], img_filename)
                if os.path.exists(img_filepath):
                    try:
                        os.remove(img_filepath)
                    except Exception:
                        pass

            db.session.delete(qr)
            deleted_count += 1
            
        log = SystemLog(
            action='all_qrs_deleted',
            details=f"Admin '{current_user.username}' deleted all {deleted_count} QR Clues.",
            user_id=current_user.id
        )
        db.session.add(log)
        db.session.commit()
        
        flash(f"Successfully deleted all {deleted_count} QR Code clues.", "success")
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting all QRs: {str(e)}")
        flash("An error occurred during bulk deletion.", "danger")

    return redirect(url_for('admin.manage_qrs'))

@admin_bp.route('/admin/qrs/delete-selected', methods=['POST'])
@check_admin_role
def delete_selected_qrs():
    import json
    ids_str = request.form.get('ids')
    if not ids_str:
        flash("No QR Codes selected.", "danger")
        return redirect(url_for('admin.manage_qrs'))
        
    try:
        ids = json.loads(ids_str)
    except Exception:
        flash("Invalid request data.", "danger")
        return redirect(url_for('admin.manage_qrs'))
        
    if not ids:
        flash("No QR Codes selected.", "danger")
        return redirect(url_for('admin.manage_qrs'))
        
    deleted_count = 0
    try:
        for qr_id in ids:
            qr = QRCode.query.get(int(qr_id))
            if qr:
                filename = f"qr_{qr.uuid}.png"
                filepath = os.path.join(current_app.config['QR_FOLDER'], filename)
                if os.path.exists(filepath):
                    try:
                        os.remove(filepath)
                    except Exception:
                        pass
                    
                if qr.image_path:
                    img_filename = qr.image_path.split('/')[-1]
                    img_filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], img_filename)
                    if os.path.exists(img_filepath):
                        try:
                            os.remove(img_filepath)
                        except Exception:
                            pass

                db.session.delete(qr)
                deleted_count += 1
                
        log = SystemLog(
            action='selected_qrs_deleted',
            details=f"Admin '{current_user.username}' deleted {deleted_count} selected QR Clues.",
            user_id=current_user.id
        )
        db.session.add(log)
        db.session.commit()
        
        flash(f"Successfully deleted {deleted_count} selected QR Code clues.", "success")
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting selected QRs: {str(e)}")
        flash("An error occurred during selection deletion.", "danger")

    return redirect(url_for('admin.manage_qrs'))

@admin_bp.route('/admin/qrs/export-zip')
@check_admin_role
def export_qrs_zip():
    """Generates all QR code PNGs and exports them as a single ZIP archive."""
    qrs = QRCode.query.filter((QRCode.clue_number > 1) | (QRCode.is_dummy == True)).all()
    qr_dir = current_app.config['QR_FOLDER']
    
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
        for qr in qrs:
            filename = f"qr_{qr.uuid}.png"
            filepath = os.path.join(qr_dir, filename)
            
            # Re-generate if file missing on disk
            if not os.path.exists(filepath):
                generate_qr_image(qr.uuid)
                
            # Define label for ZIP
            label = f"Clue_{qr.clue_number}_{qr.uuid[:6]}.png" if not qr.is_dummy else f"Dummy_{qr.uuid[:6]}.png"
            zip_file.write(filepath, label)
            
    zip_buffer.seek(0)
    
    # Audit log
    log = SystemLog(action='export_qrs_zip', details="Admin exported all QR codes as ZIP archive.", user_id=current_user.id)
    db.session.add(log)
    db.session.commit()
    
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name='treasure_hunt_qrs.zip'
    )

@admin_bp.route('/admin/qrs/export-pdf')
@check_admin_role
def export_qrs_pdf():
    """Generates a printable PDF document with QR code sheets containing clue numbers, details, and QR codes."""
    qrs = QRCode.query.filter((QRCode.clue_number > 1) | (QRCode.is_dummy == True)).order_by(QRCode.is_dummy, QRCode.clue_number).all()
    qr_dir = current_app.config['QR_FOLDER']
    
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, leftMargin=0.5*inch, rightMargin=0.5*inch, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=28,
        leading=34,
        textColor=colors.HexColor('#000000'),
        alignment=1, # Center
        spaceAfter=0
    )
    
    elements = []
    
    for idx, qr in enumerate(qrs):
        filename = f"qr_{qr.uuid}.png"
        filepath = os.path.join(qr_dir, filename)
        if not os.path.exists(filepath):
            generate_qr_image(qr.uuid)
            
        elements.append(Paragraph("LRDC 2026 TREASURE HUNT", title_style))
        elements.append(Spacer(1, 0.4*inch))
        
        # Giant QR Code Image centered (6.5 inches by 6.5 inches)
        qr_image = Image(filepath, width=6.5*inch, height=6.5*inch)
        qr_table = Table([[qr_image]], colWidths=[7.5*inch])
        qr_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 0)
        ]))
        elements.append(qr_table)
        
        # Spacer to push footer to bottom
        elements.append(Spacer(1, 1.5*inch))
        
        # Tiny bottom right identification text (< 5 pt)
        id_style = ParagraphStyle(
            f'ClueID_{idx}',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=4.5,
            leading=5,
            alignment=2, # Right
            textColor=colors.HexColor('#777777')
        )
        if qr.is_dummy:
            id_text = f"Decoy - {qr.uuid}"
        elif qr.allowed_houses:
            id_text = f"Clue #{qr.clue_number} ({qr.allowed_houses}) - {qr.uuid}"
        else:
            id_text = f"Clue #{qr.clue_number} - {qr.uuid}"
            
        id_table = Table([[Paragraph(id_text, id_style)]], colWidths=[7.5*inch])
        id_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 0)
        ]))
        elements.append(id_table)
        
        # Page break except for the last clue sheet
        if idx < len(qrs) - 1:
            elements.append(PageBreak())
            
    doc.build(elements)
    pdf_buffer.seek(0)
    
    # Audit log
    log = SystemLog(action='export_qrs_pdf', details="Admin exported all QR code clue sheets as printable PDF.", user_id=current_user.id)
    db.session.add(log)
    db.session.commit()
    
    return send_file(
        pdf_buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name='treasure_hunt_qr_sheets.pdf'
    )

@admin_bp.route('/admin/qrs/export-pdf-landscape')
@check_admin_role
def export_qrs_pdf_landscape():
    """Generates a printable PDF in Landscape orientation with 2 QR codes per page side-by-side."""
    from reportlab.lib.pagesizes import landscape
    qrs = QRCode.query.filter((QRCode.clue_number > 1) | (QRCode.is_dummy == True)).order_by(QRCode.is_dummy, QRCode.clue_number).all()
    qr_dir = current_app.config['QR_FOLDER']
    
    pdf_buffer = io.BytesIO()
    # Landscape letter: 11 x 8.5 inches
    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=landscape(letter),
        leftMargin=0.4*inch,
        rightMargin=0.4*inch,
        topMargin=0.4*inch,
        bottomMargin=0.4*inch
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitleLandscape',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=16,
        leading=20,
        textColor=colors.HexColor('#000000'),
        alignment=1, # Center
        spaceAfter=0
    )
    
    elements = []
    
    def build_qr_cell(qr, idx):
        if not qr:
            return [Spacer(1, 1)]
        
        filename = f"qr_{qr.uuid}.png"
        filepath = os.path.join(qr_dir, filename)
        if not os.path.exists(filepath):
            generate_qr_image(qr.uuid)
            
        cell_elements = []
        cell_elements.append(Paragraph("LRDC 2026 TREASURE HUNT", title_style))
        cell_elements.append(Spacer(1, 0.2*inch))
        
        # Size of QR: ~4.0 inches square (takes up most of the vertical landscape space)
        qr_image = Image(filepath, width=4.0*inch, height=4.0*inch)
        qr_image.hAlign = 'CENTER'
        cell_elements.append(qr_image)
        cell_elements.append(Spacer(1, 0.3*inch))
        
        # Identification
        id_style = ParagraphStyle(
            f'ClueID_Land_{idx}',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=4.5,
            leading=5,
            alignment=2, # Right
            textColor=colors.HexColor('#777777')
        )
        if qr.is_dummy:
            id_text = f"Decoy - {qr.uuid}"
        elif qr.allowed_houses:
            id_text = f"Clue #{qr.clue_number} ({qr.allowed_houses}) - {qr.uuid}"
        else:
            id_text = f"Clue #{qr.clue_number} - {qr.uuid}"
        
        id_p = Paragraph(id_text, id_style)
        id_p.hAlign = 'RIGHT'
        cell_elements.append(id_p)
        return cell_elements
    
    for i in range(0, len(qrs), 2):
        qr1 = qrs[i]
        qr2 = qrs[i+1] if i+1 < len(qrs) else None
        
        flow1 = build_qr_cell(qr1, i)
        flow2 = build_qr_cell(qr2, i+1)
        
        # Column widths: 5.0 inches each
        t = Table([[flow1, flow2]], colWidths=[5.0*inch, 5.0*inch])
        t.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('LEFTPADDING', (0,0), (-1,-1), 10),
            ('RIGHTPADDING', (0,0), (-1,-1), 10),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ]))
        elements.append(t)
        
        if i + 2 < len(qrs):
            elements.append(PageBreak())
            
    doc.build(elements)
    pdf_buffer.seek(0)
    
    log = SystemLog(action='export_qrs_pdf_landscape', details="Admin exported all QR codes as Landscape PDF (2 per page).", user_id=current_user.id)
    db.session.add(log)
    db.session.commit()
    
    return send_file(
        pdf_buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name='treasure_hunt_qr_sheets_landscape.pdf'
    )

@admin_bp.route('/admin/logs')
@check_admin_role
def view_logs():
    page = request.args.get('page', 1, type=int)
    action_filter = request.args.get('action_filter', '').strip()
    search = request.args.get('search', '').strip()
    
    query = SystemLog.query
    
    if action_filter:
        query = query.filter_by(action=action_filter)
        
    if search:
        query = query.filter(
            (SystemLog.details.like(f"%{search}%")) |
            (SystemLog.action.like(f"%{search}%"))
        )
        
    pagination = query.order_by(SystemLog.timestamp.desc()).paginate(page=page, per_page=25)
    logs = pagination.items
    
    # Retrieve all unique actions for filtering dropdown
    distinct_actions = db.session.query(SystemLog.action).distinct().all()
    actions = [a[0] for a in distinct_actions]

    return render_template('admin/logs.html',
                           logs=logs,
                           pagination=pagination,
                           actions=actions,
                           current_action=action_filter,
                           search=search)

@admin_bp.route('/admin/managers', methods=['GET', 'POST'])
@check_admin_role
def manage_managers():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        house_id = int(request.form.get('house_id', 0))
        
        if not username or not password or not house_id:
            flash("All fields are required.", "danger")
            return redirect(url_for('admin.manage_managers'))
            
        if User.query.filter_by(username=username).first():
            flash(f"Username '{username}' is already in use.", "danger")
            return redirect(url_for('admin.manage_managers'))

        try:
            # Create base user
            user = User(username=username, role='manager')
            user.set_password(password)
            db.session.add(user)
            db.session.flush()
            
            # Create manager profile
            mgr = Manager(user_id=user.id, house_id=house_id)
            db.session.add(mgr)
            
            # Audit log
            log = SystemLog(
                action='manager_created',
                details=f"Admin '{current_user.username}' created Manager '{username}' for House ID {house_id}.",
                user_id=current_user.id
            )
            db.session.add(log)
            db.session.commit()
            
            flash(f"Manager '{username}' created successfully.", "success")
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error creating manager: {str(e)}")
            flash("An error occurred.", "danger")
            
        return redirect(url_for('admin.manage_managers'))

    managers = Manager.query.all()
    houses = House.query.all()
    return render_template('admin/managers.html', managers=managers, houses=houses)

@admin_bp.route('/admin/managers/delete/<int:mgr_id>', methods=['POST'])
@check_admin_role
def delete_manager(mgr_id):
    mgr = Manager.query.get_or_404(mgr_id)
    user = mgr.user
    
    # Restrict deletion of default seeded managers to prevent accidental lockouts
    seeded_managers = ['Aditya', 'Shyam', 'Rahul', 'Saranya']
    if user.username in seeded_managers:
        flash(f"Cannot delete default seeded manager '{user.username}' to maintain system safety.", "warning")
        return redirect(url_for('admin.manage_managers'))

    try:
        # Delete user (deletes manager profile via cascade)
        db.session.delete(user)
        
        # Log action
        log = SystemLog(
            action='manager_deleted',
            details=f"Admin '{current_user.username}' deleted Manager '{user.username}'.",
            user_id=current_user.id
        )
        db.session.add(log)
        db.session.commit()
        
        flash("Manager deleted successfully.", "success")
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting manager: {str(e)}")
        flash("An error occurred during deletion.", "danger")

    return redirect(url_for('admin.manage_managers'))


@admin_bp.route('/admin/managers/edit/<int:mgr_id>', methods=['POST'])
@check_admin_role
def update_manager_profile(mgr_id):
    from werkzeug.security import generate_password_hash
    mgr = Manager.query.get_or_404(mgr_id)
    user = mgr.user
    
    username = request.form.get('username')
    password = request.form.get('password')
    house_id = request.form.get('house_id', type=int)
    
    if not username or not house_id:
        flash("Username and House Assignment are required.", "danger")
        return redirect(url_for('admin.manage_managers'))
        
    existing_user = User.query.filter_by(username=username).first()
    if existing_user and existing_user.id != user.id:
        flash(f"Username '{username}' is already taken.", "danger")
        return redirect(url_for('admin.manage_managers'))
        
    try:
        user.username = username
        if password and password.strip():
            user.password = generate_password_hash(password.strip())
            
        mgr.house_id = house_id
        
        log = SystemLog(
            action='manager_edited',
            details=f"Admin '{current_user.username}' edited Manager '{username}'.",
            user_id=current_user.id
        )
        db.session.add(log)
        db.session.commit()
        
        flash(f"Manager '{username}' updated successfully.", "success")
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error editing manager: {str(e)}")
        flash("An error occurred while updating the manager.", "danger")
        
    return redirect(url_for('admin.manage_managers'))


# ==========================================
# REPORT EXPORT ENDPOINTS
# ==========================================

@admin_bp.route('/admin/reports/download')
@check_admin_role
def download_reports_dashboard():
    return render_template('admin/reports.html')

@admin_bp.route('/admin/reports/export/<report_type>/<format_type>')
@check_admin_role
def export_report(report_type, format_type):
    """
    Export reports:
    - report_type: 'house_summary', 'leaderboard', 'qr_analytics', 'team_progress'
    - format_type: 'csv', 'excel', 'pdf'
    """
    timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # 1. Gather Data
    headers, data = get_report_data(report_type)
    
    if not headers:
        flash("Invalid report type.", "danger")
        return redirect(url_for('admin.download_reports_dashboard'))
        
    filename = f"{report_type}_report_{timestamp_str}.{format_type}"

    # 2. Output Formatting
    if format_type == 'csv':
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(data)
        
        output.seek(0)
        
        log = SystemLog(action=f'export_{report_type}_csv', details=f"Admin exported {report_type} as CSV.", user_id=current_user.id)
        db.session.add(log)
        db.session.commit()
        
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )
        
    elif format_type == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = report_type.replace('_', ' ').title()
        
        # Styles
        header_fill = PatternFill(start_color='0d6efd', end_color='0d6efd', fill_type='solid')
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        data_font = Font(name='Arial', size=10)
        border_side = Side(style='thin', color='DDDDDD')
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        # Write headers
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
        # Write data
        for row in data:
            ws.append(row)
            
        # Style rows and auto-fit columns
        for row_idx in range(2, len(data) + 2):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = border
                
        # Auto-adjust column width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        ws.row_dimensions[1].height = 25
        
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        log = SystemLog(action=f'export_{report_type}_excel', details=f"Admin exported {report_type} as Excel.", user_id=current_user.id)
        db.session.add(log)
        db.session.commit()
        
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    elif format_type == 'pdf':
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, leftMargin=0.5*inch, rightMargin=0.5*inch, topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=18,
            textColor=colors.HexColor('#0d6efd'),
            alignment=1,
            spaceAfter=15
        )
        
        elements = []
        elements.append(Paragraph(report_type.replace('_', ' ').upper() + " REPORT", title_style))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))
        
        # Build Table
        table_content = [headers] + [[str(val) for val in row] for row in data]
        
        # Estimate column widths based on page width (7.5 inches printable)
        num_cols = len(headers)
        if report_type == 'leaderboard':
            col_widths = [0.5*inch, 1.4*inch, 0.8*inch, 0.6*inch, 1.0*inch, 1.0*inch, 2.2*inch]
            t = Table(table_content, colWidths=col_widths)
        else:
            col_width = (7.5 * inch) / num_cols
            t = Table(table_content, colWidths=[col_width]*num_cols)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0d6efd')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('TOPPADDING', (0,0), (-1,0), 8),
            ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#f8f9fa')),
            ('TEXTCOLOR', (0,1), (-1,-1), colors.black),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,1), (-1,-1), 9),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dee2e6')),
            ('PADDING', (0,0), (-1,-1), 5),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
        ]))
        
        elements.append(t)
        doc.build(elements)
        pdf_buffer.seek(0)
        
        log = SystemLog(action=f'export_{report_type}_pdf', details=f"Admin exported {report_type} as PDF.", user_id=current_user.id)
        db.session.add(log)
        db.session.commit()
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    flash("Invalid format type.", "danger")
    return redirect(url_for('admin.download_reports_dashboard'))

def get_report_data(report_type):
    """Data helper mapping report queries to output lists."""
    if report_type == 'house_summary':
        headers = ['House Name', 'Total Teams', 'R1 Pending', 'R1 Completed', 'R2 Started', 'R2 Finished', 'Total QR Scans']
        data = []
        houses = House.query.all()
        for h in houses:
            teams = h.teams
            t_ids = [t.id for t in teams]
            
            r1_pend = Team.query.filter_by(house_id=h.id, round1_status='requested').count()
            r1_comp = Team.query.filter(Team.house_id == h.id, Team.round1_status.in_(['verified', 'approved'])).count()
            r2_start = Team.query.filter_by(house_id=h.id, current_round=2, round2_completed=False).count()
            r2_fin = Team.query.filter_by(house_id=h.id, round2_completed=True).count()
            
            scans = 0
            if t_ids:
                scans = QRScanLog.query.filter(QRScanLog.team_id.in_(t_ids)).count()
                
            data.append([
                h.name,
                len(teams),
                r1_pend,
                r1_comp,
                r2_start,
                r2_fin,
                scans
            ])
        return headers, data
        
    elif report_type == 'leaderboard':
        headers = ['Rank', 'Team Name', 'House', 'Points', 'Clue Progress', 'Total Time', 'Clue Durations']
        data = []
        
        teams = Team.query.all()
        
        def sort_key(t):
            solved_count = 7 if t.round2_completed else (t.round2_current_clue - 1 if t.current_round == 2 else 0)
            points = (100 if t.round1_status == 'approved' else 0) + solved_count * 10
            
            duration = 0
            r1_app = Round1Approval.query.filter_by(team_id=t.id).first()
            r2_start = r1_app.approved_at if (r1_app and r1_app.approved_at) else t.created_at
            
            if t.round2_completed:
                duration = (t.round2_completion_time - r2_start).total_seconds() if (t.round2_completion_time and r2_start) else 0
            elif t.current_round == 2:
                clues_solved = t.round2_current_clue - 1
                if clues_solved == 1:
                    duration = 0
                elif clues_solved > 1:
                    prog = Round2Progress.query.filter_by(team_id=t.id, clue_number=clues_solved).first()
                    if prog and r2_start:
                        duration = (prog.completed_at - r2_start).total_seconds()
            else:
                comp_count = len(t.task_completions)
                if comp_count > 0:
                    last_task = TaskCompletion.query.filter_by(team_id=t.id).order_by(TaskCompletion.completed_at.desc()).first()
                    if last_task:
                        duration = (last_task.completed_at - t.created_at).total_seconds()
            return (-points, duration)
            
        sorted_teams = sorted(teams, key=sort_key)
        
        for rank, t in enumerate(sorted_teams, start=1):
            solved_count = 7 if t.round2_completed else (t.round2_current_clue - 1 if t.current_round == 2 else 0)
            points = (100 if t.round1_status == 'approved' else 0) + solved_count * 10
            
            r1_app = Round1Approval.query.filter_by(team_id=t.id).first()
            r2_start = r1_app.approved_at if (r1_app and r1_app.approved_at) else t.created_at
            
            progresses = Round2Progress.query.filter_by(team_id=t.id).order_by(Round2Progress.clue_number).all()
            prog_map = {p.clue_number: p.completed_at for p in progresses}
            if t.current_round == 2:
                prog_map[1] = r2_start
            
            # Duration shows total active elapsed time since Round 2 started
            total_dur_sec = 0
            if t.round2_completed:
                total_dur_sec = int((t.round2_completion_time - r2_start).total_seconds()) if (t.round2_completion_time and r2_start) else 0
            elif t.current_round == 2:
                total_dur_sec = int((datetime.utcnow() - r2_start).total_seconds()) if r2_start else 0
            else:
                total_dur_sec = 0

            durations_text_list = []
            for lvl in range(1, 8):
                if lvl in prog_map:
                    if lvl == 1:
                        dur_sec = int((prog_map[1] - r2_start).total_seconds())
                    else:
                        prev_time = prog_map.get(lvl - 1)
                        if prev_time:
                            dur_sec = int((prog_map[lvl] - prev_time).total_seconds())
                        else:
                            dur_sec = 0
                    m = dur_sec // 60
                    s = dur_sec % 60
                    durations_text_list.append(f"C{lvl}:{m}m{s}s")
                else:
                    if t.current_round == 2 and t.round2_current_clue == lvl:
                        prev_time = prog_map.get(lvl - 1)
                        if prev_time:
                            active_dur = int((datetime.utcnow() - prev_time).total_seconds())
                        else:
                            active_dur = 0
                        m = active_dur // 60
                        s = active_dur % 60
                        durations_text_list.append(f"C{lvl}:{m}m{s}s*")
                    else:
                        durations_text_list.append(f"C{lvl}:-")
                    
            h_tot = total_dur_sec // 3600
            m_tot = (total_dur_sec % 3600) // 60
            s_tot = total_dur_sec % 60
            total_time_str = f"{h_tot:02d}:{m_tot:02d}:{s_tot:02d}" if total_dur_sec > 0 else "00:00:00"
            if t.round1_status != 'approved' and t.current_round == 1:
                total_time_str = "-"
                
            clue_durs_str = ", ".join(durations_text_list)
            qr_prog = f"{solved_count} / 7 solved" if t.current_round == 2 or t.round2_completed else "Not started"
            
            data.append([
                rank,
                t.team_name,
                t.house.name,
                points,
                qr_prog,
                total_time_str,
                clue_durs_str,
                t.winner_rank
            ])
        return headers, data
        
    elif report_type == 'qr_analytics':
        headers = ['Timestamp', 'Team Name', 'House', 'Clue Target', 'Scan Type', 'IP Address', 'Browser']
        data = []
        logs = QRScanLog.query.order_by(QRScanLog.timestamp.desc()).all()
        for log in logs:
            scan_type = "Correct"
            if log.is_dummy:
                scan_type = "Dummy Clue"
            elif log.is_repeated:
                scan_type = "Repeated Scan"
            elif not log.qr_code_id:
                scan_type = "Invalid Token"
            elif not log.is_correct:
                scan_type = "Out of Order"
                
            clue_tgt = f"Clue {log.qr_code.clue_number}" if log.qr_code else "None"
            if log.is_dummy:
                clue_tgt = "Dummy Clue"
                
            data.append([
                log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                log.team.team_name,
                log.team.house.name,
                clue_tgt,
                scan_type,
                log.ip_address or "-",
                log.browser[:30] + "..." if log.browser else "-"
            ])
        return headers, data
        
    elif report_type == 'team_progress':
        headers = ['Team Name', 'House', 'Leader Phone', 'R1 Tasks Done', 'R1 Status', 'R2 Progress', 'Last Activity']
        data = []
        teams = Team.query.all()
        for t in teams:
            r1_done = f"{len(t.task_completions)} / 10"
            r2_prog = f"{t.round2_current_clue - 1} / 7" if t.current_round == 2 else "Locked"
            if t.round2_completed:
                r2_prog = "Completed"
                
            # Find last activity time
            last_act = "-"
            last_scan = QRScanLog.query.filter_by(team_id=t.id).order_by(QRScanLog.timestamp.desc()).first()
            last_task = TaskCompletion.query.filter_by(team_id=t.id).order_by(TaskCompletion.completed_at.desc()).first()
            
            times = []
            if last_scan: times.append(last_scan.timestamp)
            if last_task: times.append(last_task.completed_at)
            
            if times:
                last_act = max(times).strftime('%Y-%m-%d %H:%M:%S')
            else:
                last_act = t.created_at.strftime('%Y-%m-%d %H:%M:%S')
                
            data.append([
                t.team_name,
                t.house.name,
                t.leader_phone,
                r1_done,
                t.round1_status.upper(),
                r2_prog,
                last_act
            ])
        return headers, data
        
    return None, None
